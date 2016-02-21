# -*- coding: utf-8 -*-
"""
Created on Fri Nov 27 11:06:25 2015

@author: Alex
"""

import xml.etree.ElementTree as ET
from openpyxl import Workbook
import platform
if platform.system() == 'Windows':
    from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
    import openpyxl.utils as Utils
import sqlite3 as lite
import os
from ObjectClasses import _Workflow, _KeyAction, _InputParameter

class TemplateReader():
    #Read the XML Template and generate SQL Queries based on it
    #Pass to the writer each segment individually

    def __init__(self, db_path):
        self.wb = Workbook()
        orig_sheet = self.wb.get_active_sheet()
        if platform.system() == 'Windows':
            self.wb.create_sheet('Header', 0)
        else:
            self.wb.create_sheet(0, 'Header')
        self.wb.remove_sheet(orig_sheet)
        self.db_path=db_path
        
        if platform.system() == 'Windows':
        
            #Base Styles
            self.base_font = Font(name='Calibri', size=11, bold=False, italic=False, vertAlign=None, underline='none', strike=False, color='FF000000')
            self.base_fill = PatternFill(fill_type=None, start_color='FFFFFFFF', end_color='FF000000')
            self.base_border = Border(left=Side(border_style=None, color='FF000000'), right=Side(border_style=None, color='FF000000'),\
                top=Side(border_style=None, color='FF000000'), bottom=Side(border_style=None, color='FF000000'), diagonal=Side(border_style=None, color='FF000000'),\
                    diagonal_direction=0, outline=Side(border_style=None, color='FF000000'), vertical=Side(border_style=None, color='FF000000'), horizontal=Side(border_style=None, color='FF000000'))
            self.base_alignment=Alignment(horizontal='general', vertical='bottom', text_rotation=0, wrap_text=True, shrink_to_fit=False, indent=0)
            self.base_number_format = 'General'
            
            #Base Header Styles
            self.header_font = Font(name='Calibri', size=13, bold=True, italic=False, vertAlign=None, underline='none', strike=False, color='FF000000')
            self.header_fill = PatternFill(fill_type=None, start_color='FFFFFFFF', end_color='FF000000')
            self.header_border = Border(left=Side(border_style=None, color='FF000000'), right=Side(border_style=None, color='FF000000'),\
                top=Side(border_style=None, color='FF000000'), bottom=Side(border_style=None, color='FF000000'), diagonal=Side(border_style=None, color='FF000000'),\
                    diagonal_direction=0, outline=Side(border_style='thin', color='FF000000'), vertical=Side(border_style=None, color='FF000000'), horizontal=Side(border_style=None, color='FF000000'))
            self.header_alignment=Alignment(horizontal='center', vertical='center', text_rotation=0, wrap_text=True, shrink_to_fit=False, indent=0)
            self.header_number_format = 'General'

    def select_files_in_folder(self, dir, ext):
        for file in os.listdir(dir):
            if file.endswith('.%s' % ext):
                yield os.path.join(dir, file)
                
    def generate_parameter_list(self, xml_path):
        
        #Param list to hold input parameters
        param_list=[]
        #Read the XML Template
        self.tree = ET.parse(xml_path)
        self.root = self.tree.getroot()
        for child in self.root:
            if child.tag == 'InputParameters':
                for value in child:
                    #Add each value to the parameter list
                    param_list.append(value.text)
        return param_list
        
    #Write a Key Action Block
    #Pass in a key action object, and the row and column
    def write_key_action_block(self, keyaction, int_row, int_col):
        #Write the action name
        ka_cell = '%s%s' % (Utils.get_column_letter(int_col), int_row)
        worksheet[ka_cell] = action.name
#            worksheet.merge_cells('%s%s:%s%s' % (Utils.get_column_letter(int_col), int_row, Utils.get_column_letter(int_col), int_row + action.numParams()))
        
        #Write the action description
        desc_cell = '%s%s' % (Utils.get_column_letter(int_col+1), int_row)
        worksheet[desc_cell] = action.description
#            worksheet.merge_cells('%s%s:%s%s' % (Utils.get_column_letter(int_col+1), int_row, Utils.get_column_letter(int_col + 1), int_row + action.numParams()))
        
        #Write the input parameters for the key action
        param_row = int_row
        for param in action.input_parameters:
            param_cell = '%s%s' % (Utils.get_column_letter(int_col+2), param_row)
            worksheet[param_cell] = param.name
            param_value_cell = '%s%s' % (Utils.get_column_letter(int_col+3), param_row)
            worksheet[param_value_cell] = param.value
            param_row+=1
            
        #Write Expected Result
        er_cell = '%s%s' % (Utils.get_column_letter(int_col+4), int_row)
        worksheet[er_cell] = action.expected_result
#            worksheet.merge_cells('%s%s:%s%s' % (Utils.get_column_letter(int_col+4), int_row, Utils.get_column_letter(int_col + 1), int_row + action.numParams()))
        if action.numParams() == 0:
            int_row+=1
        else:
            int_row+=action.numParams()
                
    #Take a _Workflow object and write it to the specified excel worksheet
    #This is a very basic export without ordering, should be removed
    def execute_workflow_export(self, flow, worksheet, row, column):
        #Export the _Workflow object to the specified row and column on the excel sheet
        
        int_row = row
        int_col = column
        num_columns=9
        
        #Write the top header
        top_cell = '%s%s' % (Utils.get_column_letter(int_col), int_row)
        worksheet[top_cell] = flow.name
        if platform.system() == 'Windows':
            worksheet[top_cell].font = self.header_font
            worksheet[top_cell].fill = self.header_fill
            worksheet[top_cell].border = self.header_border
            worksheet[top_cell].alignment = self.header_alignment
            worksheet[top_cell].number_format = self.header_number_format
        worksheet.merge_cells('%s%s:%s%s' % (Utils.get_column_letter(int_col), int_row, Utils.get_column_letter(int_col+num_columns), int_row))
        int_row+=1
        
        #Write the column headers
        column_list = ['Workflow', 'Action', 'Description', 'Parameter', 'Value', 'Expected Result', 'Notes', 'Pass/Fail', "Tester's Initials"]
        
        list_counter=0
        for column in column_list:
            cell = '%s%s' % (Utils.get_column_letter(int_col), int_row)
            worksheet[cell] = column_list[list_counter]
            int_col+=1
            list_counter+=1
            
            if platform.system() == 'Windows':
                worksheet[cell].font = self.header_font
                worksheet[cell].fill = self.header_fill
                worksheet[cell].border = self.header_border
                worksheet[cell].alignment = self.header_alignment
                worksheet[cell].number_format = self.header_number_format
        int_col=1
        int_row+=1
        
        #Write the workflow
        wf_cell = '%s%s' % (Utils.get_column_letter(int_col), int_row)
        worksheet[wf_cell] = flow.name
        if flow.numRows() == 0:
            numrows = 0
        else:
            numrows = flow.numRows() - 1
        worksheet.merge_cells('%s%s:%s%s' % (Utils.get_column_letter(int_col), int_row, Utils.get_column_letter(int_col), int_row + numrows))
        int_col = 2
        
        #Write the KeyActions
        for action in flow.keyactions:
            
            #Write the action name
            ka_cell = '%s%s' % (Utils.get_column_letter(int_col), int_row)
            worksheet[ka_cell] = action.name
#            worksheet.merge_cells('%s%s:%s%s' % (Utils.get_column_letter(int_col), int_row, Utils.get_column_letter(int_col), int_row + action.numParams()))
            
            #Write the action description
            desc_cell = '%s%s' % (Utils.get_column_letter(int_col+1), int_row)
            worksheet[desc_cell] = action.description
#            worksheet.merge_cells('%s%s:%s%s' % (Utils.get_column_letter(int_col+1), int_row, Utils.get_column_letter(int_col + 1), int_row + action.numParams()))
            
            #Write the input parameters for the key action
            param_row = int_row
            for param in action.input_parameters:
                param_cell = '%s%s' % (Utils.get_column_letter(int_col+2), param_row)
                worksheet[param_cell] = param.name
                param_value_cell = '%s%s' % (Utils.get_column_letter(int_col+3), param_row)
                worksheet[param_value_cell] = param.value
                param_row+=1
                
            #Write Expected Result
            er_cell = '%s%s' % (Utils.get_column_letter(int_col+4), int_row)
            worksheet[er_cell] = action.expected_result
#            worksheet.merge_cells('%s%s:%s%s' % (Utils.get_column_letter(int_col+4), int_row, Utils.get_column_letter(int_col + 1), int_row + action.numParams()))
            if action.numParams() == 0:
                int_row+=1
            else:
                int_row+=action.numParams()
            
    #Query the DB And generate a _Workflow object
    def generate_workflow_export(self, workflow_id):
        
        #The number of rows in the workflow to be returned
        num_rows = 0
        
        #Find the workflow
        self.cur.execute("select wf.id from workflow wf where wf.id = %s" % (workflow_id))
#        self.cur.execute("select wf.id, wf.name from workflow wf left join testscript ts on ts.id = wf.testscriptid left join project p on ts.projectid = p.id left join client c on p.clientid = c.id where wf.name = '%s' and ts.name = '%s' and p.name = '%s' and c.name = '%s' order by wf.id;" % (workflow_name, testscript, project, client))
        flow = self.cur.fetchone()
        
        workflow = _Workflow()
        workflow.name = flow[1]
        workflow.id = flow[0]
        print('Workflow %s created with ID %s' % (workflow.name, workflow.id))
        
        #Find the Key Actions for the workflow
        self.cur.execute("select ka.id, ka.name, ka.description, ka.custom, wfa.expectedresult, wfa.notes from workflowaction wfa left join keyaction ka on wfa.keyactionid = ka.id left join workflow w on wfa.workflowid = w.id where w.id = '%s';" % (workflow.id))
        keyactions = self.cur.fetchall()
        
        for action in keyactions:
            
            keyaction = _KeyAction()
            keyaction.id = action[0]
            keyaction.name = action[1]
            keyaction.description = action[2]
            if action[3] == 0:
                keyaction.custom = False
            else:
                keyaction.custom = True
            keyaction.expected_result = action[4]
            keyaction.notes = action[5]
            print('Key Action %s created' % (keyaction.name))
            print('Description: %s' % (keyaction.description))
            print('Custom: %s' % (keyaction.custom))
            print('Expected Result: %s' % (keyaction.expected_result))
            
            #Find the Next Actions for the Key Action
            self.cur.execute("select ka.id, ka.name, ka.description, ka.custom, wfa.expectedresult, wfa.notes from workflowaction wfa left join keyaction ka on wfa.keyactionid = ka.id where ka.id = %s;" % (action.id))
            keyactions = self.cur.fetchall()
            for action in nextactions:
            
                ka = _KeyAction()
                ka.id = action[0]
                ka.name = action[1]
                ka.description = action[2]
                if action[3] == 0:
                    ka.custom = False
                else:
                    ka.custom = True
                ka.expected_result = action[4]
                ka.notes = action[5]
                print('Key Action %s created' % (ka.name))
                print('Description: %s' % (ka.description))
                print('Custom: %s' % (ka.custom))
                print('Expected Result: %s' % (ka.expected_result))
                keyaction.add_nextaction(ka)
            
            #Find the Input Parameters for the Key Action
            self.cur.execute('select ip.id, ip.name, wp.value from ((inputparameter ip left join keyaction ka on ip.keyactionid = ka.id) left join workflowparam wp on wp.inputparamid = ip.id) where ka.id = %s;' % (action[0]))
            inputparameters = self.cur.fetchall()
            if len(inputparameters) == 0:
                num_rows+=1
            else:
                for param in inputparameters:
                    input_parameter = _InputParameter()
                    input_parameter.id = param[0]
                    input_parameter.name = param[1]
                    input_parameter.value = param[2]
                    keyaction.add_inputparameter(input_parameter)
                    print('Input Parameter %s added to keyaction %s' % (input_parameter.name, keyaction.name))
                    num_rows+=1
            workflow.add_keyaction(keyaction)
            print('Key Action %s added to Workflow %s' % (keyaction.name, workflow.name))
                
        #Write the _Workflow object to the Excel Sheet
#        self.execute_workflow_export(workflow, worksheet, row, column)
            
        return workflow
        
    def process_for_wildcard(self, text, params, wc_counter, param_counter):
        wc_counter = 0
        for i in text:
            if i == '?':
                text[wc_counter] = params[int(text[wc_counter + 1])]
                print('Parameter %s used' % (params[int(text[wc_counter + 1])]))
                del text[wc_counter + 1]
                param_counter+=1
            wc_counter+=1
            
    def set_header_font(self, cell):
        if platform.system() == 'Windows':
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.header_border
            cell.alignment = self.header_alignment
            cell.number_format = self.header_number_format
            
    def set_base_font(self, cell):
        if platform.system() == 'Windows':
            cell.font = self.base_font
            cell.fill = self.base_fill
            cell.border = self.base_border
            cell.alignment = self.base_alignment
            cell.number_format = self.base_number_format
            
    def write_workflow_header(self, row, column):
        pass

    #Takes in the XML path of the template file
    #params is a list of the input parameters for the template
    #Look for '?' in the xml values and replace it with the next up param
    def translate_template(self, xml_path, params):
        
        #Counter for the input parameters
        param_counter = 0
        
        #Counter for the segments within pages
        segment_counter = 0
        
        #Wild Card Counter
        wc_counter = 0
        
        #Clear the sheets out of the workbook
        sn = self.wb.get_sheet_names()
        for s in sn:
            self.wb.remove_sheet(self.wb.get_sheet_by_name(s))
        self.wb.create_sheet('Header', 0)
        
        #Connect to the DB
        self.con = lite.connect(self.db_path)
        self.cur = self.con.cursor()
        
        with self.con:
        
            #Read the XML Template
            self.tree = ET.parse(xml_path)
            self.root = self.tree.getroot()
            header_ws = self.wb.get_sheet_by_name('Header')
            for child in self.root:
                if child.tag == 'Header':
                    for row in child:
                        #Process each header row
                        for element in row:
                            #Process each element
                            #We expect two property values:
                            #cell, the first cell of the merge
                            #end_cell, the last cell in the merge
                            if '?' not in element.text:
                                header_ws[element.attrib['start_cell']] = element.text
                            #If a wildcard is encountered, we need to replace it with the
                            #correct parameter
                            else:
                                text = list(element.text)
                                text = self.process_for_wildcard(text, params, wc_counter, param_counter)
                                header_ws[element.attrib['start_cell']] = ''.join(text)
                            self.set_header_font(header_ws[element.attrib['start_cell']])
                            header_ws.merge_cells('%s:%s' % (element.attrib['start_cell'], element.attrib['end_cell']))
                            print('Header element placed')
                elif child.tag == 'Body':
                    #Process the body segment
                    for page in child:
                        #Process each page
                        segment_counter = 0
                        body_ws = self.wb.create_sheet(page.attrib['name'])
                        for segment in page:
                            if segment.tag == 'TestScript':
                                #Execute the specialized Test Script Steps Export
                                #This segment needs to be hardcoded due to the ability to construct
                                #nonlinear workflows.
                                #Here, we process a full testscript and output each workflow in an optimized state
                            
                                start_cell = segment.attrib['cell']
                                flow_list=[]
                                header_list=[]
                                db_column_list=[]
                                for query_segment in segment:
                                    if query_segment.tag == 'Query':
                                        #Execute the query and place the results into the flow list, 
                                        #we need the first column in the query to be workflow.id and this should be unique
                                        if '?' not in query_segment.text:
                                            self.cur.execute(query_segment.text)
                                        #If a wildcard is encountered, we need to replace it with the
                                        #correct parameter
                                        else:
                                            text = list(query_segment.text)
                                            text = self.process_for_wildcard(text, params, wc_counter, param_counter)
                                            self.cur.execute(text)
                                        print('query %s executed' % (text))
                                        workflows = self.cur.fetchall()
                                        
                                        #Find the flow list for the testscript
                                        for workflow in workflows:
                                            flow = self.generate_workflow_export(workflow[0])
                                            flow_list.append(flow)
                                            
                                    if query_segment.tag == 'KeyActionBlock':
                                        for header_segment in query_segment:
                                            if header_segment.tag == 'Column':
                                                header_list.append(header_segment.text)
                                                db_column_list.append(header_segment.text)
                                            elif header_segment.tag == 'Blank':
                                                header_list.append(header_segment.text)
                                        
                                #Find the row and col of the start cell
                                col = Utils.column_index_from_string(start_cell[0])
                                row = int(float(start_cell[1]))
                                
                                for f in flow_list:
                                    #Build the key action trees and store them in memory
                                    f.build_keyactiontree()
                                    
                                    #TO-DO: Write the workflow to the sheet
                                    
                                    #Write the header
                                    self.write_workflow_header(header_list, row, col)

                            else:
                                for child in segment:
                                    if child.tag == 'Title':
                                        if '?' not in child.text:
                                            body_ws[segment.attrib['cell']] = child.text
                                        #If a wildcard is encountered, we need to replace it with the
                                        #correct parameter
                                        else:
                                            text = list(child.text)
                                            text = self.process_for_wildcard(text, params, wc_counter, param_counter)
                                            body_ws[segment.attrib['cell']] = ''.join(text)
                                        self.set_header_font(body_ws[segment.attrib['cell']])
                                        print('Data Title element %s placed in cell %s' % (child.text, segment.attrib['cell']))
                                        segment_counter+=1
                                    elif child.tag == 'Header':
                                        i=0
                                        for column in child:
                                            #Place the column header for each query column
                                            cell = Utils.coordinate_from_string(segment.attrib['cell'])
                                            col = Utils.column_index_from_string(cell[0])
                                            body_ws['%s%s' % (Utils.get_column_letter(col+i), 1 + segment_counter)] = column.text
                                            self.set_base_font(body_ws['%s%s' % (Utils.get_column_letter(col+i), 1 + segment_counter)])
                                            print('Data Header element %s placed in cell %s%s' % (column.text, Utils.get_column_letter(col+i), 2))
                                            i+=1
                                        segment_counter+=1
                                    elif child.tag == 'Query':
                                        #Execute the query and place the results into the page
                                        if '?' not in child.text:
                                            self.cur.execute(child.text)
                                        #If a wildcard is encountered, we need to replace it with the
                                        #correct parameter
                                        else:
                                            text = list(child.text)
                                            text = self.process_for_wildcard(text, params, wc_counter, param_counter)
                                            self.cur.execute(text)
                                        data = self.cur.fetchall()
                                        print('query %s executed' % (text))
                                        i=3
                                        for row in data:
                                            j=0
                                            segment_counter+=1
                                            #Place the data into the report
                                            for e in row:
                                                cell = Utils.coordinate_from_string(segment.attrib['cell'])
                                                col = Utils.column_index_from_string(cell[0])
                                                body_ws['%s%s' % (Utils.get_column_letter(col+j), i)] = e
                                                self.set_base_font(body_ws['%s%s' % (Utils.get_column_letter(col+j), i)])
                                                print('Data Element %s placed in column %s%s' % (e, Utils.get_column_letter(col+i), j))
                                                j+=1
                                            i+=1
            self.wb.save('Export.xlsx')